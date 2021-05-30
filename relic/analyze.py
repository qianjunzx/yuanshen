from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
wb=load_workbook(filename='relic_data.xlsx')
relic_condition=[1,1,1,1,1]
    
def relic_board(relic_condition):#圣遗物面板加合，输出格式为列表，1.攻击力 2.攻击力百分比...(见antry_type字典)
    global entry_type
    outcome=[0,0,0,0,[],[],0,0,0,0,0,0,0,[],[],0]#圣遗物面板输出
    for round1 in range(1,6):
        sheet1=wb[wb.sheetnames(round1)]
        for round2 in range(2,7):
            cell1=sheet1.cell(row=relic_condition[round1-1],column=round2).value
            cell2=sheet1.cell(row=relic_condition[round1-1],column=round2+5).value
        if isdamage(cell1) != 0:
            func=outcome[5]
            func.append(cell2)
            outcome[5]=func
            func=outcome[6]
            func.append(isdamage(cell1))
            outcome[6]=func
        else:
            outcome[entry_type[cell1]] += cell2
    return outcome

entry_type={#词条和编号的映射
    '攻击力':        1,
    '攻击力百分比':   2,
    '暴击率':        3,
    '暴击伤害':      4,
    '防御力':        7,
    '防御力百分比':  8,
    '生命值':       9,
    '生命值百分比': 10,
    '治疗加成':     11,
    '元素精通':     12,
    '元素充能效率': 13,
    '护盾强效':     16,
}

def isdamage(a):#伤害加成类型和编号的映射（顺带用于判断词条是否为增伤）
    if a == '火元素伤害加成':
        return 1
    elif a == '水元素伤害加成':
        return 2
    elif a == '冰元素伤害加成':
        return 3
    elif a == '雷元素伤害加成':
        return 4
    elif a == '草元素伤害加成':
        return 5
    elif a == '风元素伤害加成':
        return 6
    elif a == '岩元素伤害加成':
        return 7
    elif a == '物理伤害加成':
        return 8
    elif a== '元素战技伤害加成':
        return 9
    elif a== '元素爆发伤害加成':
        return 10
    elif a== '普通攻击伤害加成':
        return 11
    elif a== '重击伤害加成':
        return 12
    elif a== '伤害加成':
        return 13
    else:
        return 0

suit_effort={           #确定套件的加成【两件套大类（1.普通加成2.伤害加成3.特殊处理），两天套小类，两件套数值，四件套大类，四件套小类，四件套数值】
    '冰风迷途的勇士'  : [2,3,15,1,3,40],
    '平息鸣雷的尊者'  : [0,0,0,2,4,35],
    '渡过烈火的贤人'  : [0,0,0,2,1,35],
    '被怜爱的少女'    : [1,11,15,1,11,20],
    '角斗士的终幕礼'   :[1,2,18,2,11,35],
    '翠绿之影'        : [2,6,15,3,0,0],#unfinished
    '流浪大地的乐团'   :[1,13,80,2,12,35],
    '如雷的盛怒'      : [2,4,15,3,0,0],#unfinished
    '炽烈的炎之魔女'    : [2,1,15,2,1,7.5],#unfinishied
    '昔日之宗室'        :[2,10,20,1,2,20],
    '染血的骑士'        :[2,8,25,3,0,0],#unfinished
    '悠古的磐岩'        :[2,7,15,3,0,0],#unfinished
    '逆飞的流行'        :[1,16,35,2,13,35],
    '千岩牢固'          :[1,10,20,1,2,30],#unfinished
    '苍白之火'          :[2,8,25,2,8,25,12,18],
}

def relic_suit(relic_condition):#查看套件类型
    suit={}
    suit_type=[]
    for round1 in range(0,5):
        sheet1=wb[wb.sheetnames(round1+1)]
        cell=sheet1.cell(row=relic_condition[round1],column=1).value
        suit[cell]+=1
        if suit[cell]==1:
            suit_type.append(cell)
    return suit,suit_type
def suit_condition(suit,suit_type,outcome):#
    for round1 in suit_type:
        effort_list=suit_effort[round1]
        suit_quant=suit[round1]
        if suit_quant >= 2:
            if effort_list[0]== 1:
                outcome[effort_list[1]]+=effort_list[2]
            if effort_list[0]== 2:
                outcome[5].append(effort_list[2])
                outcome[6].append(effort_list[1])
        if suit_quant >= 4:
            i=int(len(effort_list)/3)
            for round2 in range(2,i+1):
                j=3*round2-4
                if effort_list[j]== 1:
                    outcome[effort_list[j+1]]+=effort_list[j+2]
                if effort_list[0]== 2:
                    outcome[5].append(effort_list[j+2])
                    outcome[6].append(effort_list[j+1])
    return outcome

def character_board(character,level):
    pass
