from openpyxl import load_workbook
workbook=load_workbook(filename='原神祈愿记录_20210524_175903.xlsx')
sheet3 = workbook[workbook.sheetnames[3]]
my_list=[]
quant =0
out_come=[]
def init(a):
    global quant
    global my_list
    quant= quant + 1
    if not(a in my_list):
        my_list.append(a)
        return len(my_list),a,quant
for i in range(1,sheet3.max_row+1):
    if sheet3.cell(row=i,column=4).value ==3:
        out_come.append(sheet3.cell(row=i,column=2).value)
for j in out_come:
    print(j)